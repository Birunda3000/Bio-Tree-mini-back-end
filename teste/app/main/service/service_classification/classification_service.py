from math import ceil
from re import match
from warnings import filterwarnings

from openpyxl import load_workbook
from pyexcel_ods3 import get_data
from sqlalchemy import and_, or_
from werkzeug.datastructures import FileStorage, ImmutableMultiDict

from app.main import db
from app.main.config import Config
from app.main.exceptions import DefaultException, ValidationException
from app.main.model import Classification, Service

_CONTENT_PER_PAGE = Config.DEFAULT_CONTENT_PER_PAGE


def get_classifications(params: ImmutableMultiDict):

    page = params.get("page", type=int, default=1)
    per_page = params.get("per_page", type=int, default=_CONTENT_PER_PAGE)
    name = params.get("name", type=str)
    service_id = params.get("service_id", type=int)

    filters = []

    if name:
        filters.append(Classification.name.ilike(f"%{name}%"))
    if service_id:
        filters.append(Classification.service_id == service_id)

    pagination = (
        Classification.query.filter(*filters)
        .order_by(Classification.id)
        .paginate(page=page, per_page=per_page, error_out=False)
    )

    return {
        "current_page": page,
        "total_items": pagination.total,
        "total_pages": ceil(pagination.total / per_page),
        "items": pagination.items,
    }


def save_new_classification(data: dict[str, any]) -> None:

    service_id = int(data.get("service_id"))
    get_service(service_id=service_id)

    code = data.get("code")
    name = data.get("name").upper()

    filters = [
        or_(
            and_(
                Classification.service_id == service_id,
                Classification.code == code,
            ),
            and_(
                Classification.service_id == service_id,
                Classification.name == name,
            ),
        )
    ]

    _check_if_classification_already_exists(filters=filters, name=name)

    new_classification = Classification(
        service_id=service_id,
        code=code,
        name=name,
    )

    db.session.add(new_classification)
    db.session.commit()


def update_classification(classification_id: int, data: dict[str, str]) -> None:

    classification = get_classification(classification_id=classification_id)

    new_service_id = data.get("service_id")
    new_code = data.get("code")
    new_name = data.get("name").upper()

    get_service(service_id=new_service_id)

    filters = []

    if change_name := classification.name != new_name:
        filters.append(
            and_(
                Classification.service_id == new_service_id,
                Classification.code == new_code,
            )
        )

    if change_code := classification.code != new_code:
        filters.append(
            and_(
                Classification.service_id == new_service_id,
                Classification.name == new_name,
            )
        )

    if change_name or change_code:
        filters = [Classification.id != classification.id, or_(*filters)]
        _check_if_classification_already_exists(filters=filters, name=new_name)

    classification.service_id = new_service_id
    classification.name = new_name
    classification.code = new_code

    db.session.commit()


def delete_classification(classification_id: int):

    classification = get_classification(classification_id=classification_id)

    db.session.delete(classification)
    db.session.commit()


def get_classification(classification_id: int, options: list = None) -> Classification:

    query = Classification.query

    if options is not None:
        query = query.options(*options)

    classification = query.get(classification_id)

    if classification is None:
        raise DefaultException("classification_not_found", code=404)

    return classification


def _check_if_classification_already_exists(filters: list[any], name: str) -> None:
    if (
        classification_db := Classification.query.with_entities(
            Classification.service_id, Classification.name
        )
        .filter(*filters)
        .first()
    ):
        if classification_db.name == name:
            raise DefaultException("name_in_use", code=409)
        else:
            raise DefaultException("code_in_use", code=409)


def _classification_exists(
    service_id: int,
    classification_code: int,
    classification_name: str,
    classification_id: int = None,
) -> bool:

    filters = []
    if classification_id:
        filters.append(Classification.id != classification_id)

    return (
        Classification.query.filter(
            or_(
                and_(
                    Classification.service_id == service_id,
                    Classification.code == classification_code,
                ),
                and_(
                    Classification.service_id == service_id,
                    Classification.name == classification_name,
                ),
            ),
            *filters,
        )
        .with_entities(Classification.id)
        .first()
        is not None
    )


def check_numbers(string):
    pattern = r"^\d{3,3}$"
    result = match(pattern, string)
    return result


def read_ods(file: FileStorage) -> dict[str, str]:

    data = dict(get_data(file))

    services_classifications = dict()
    try:
        for line in data[list(data.keys())[0]]:
            if line and len(line) == 2 and line[1].startswith("Serviço:"):
                service = line[-1].split(": ", 1)[-1].upper()
                services_classifications.setdefault(service, [])
            elif len(line) == 3:
                services_classifications[service].append(line[-1].upper())
    except:
        raise DefaultException("non_standard_file", code=400)

    return services_classifications


def read_xlsx(file: FileStorage) -> dict[str, str]:

    filterwarnings("ignore", category=UserWarning, module="openpyxl")
    workbook = load_workbook(file, keep_vba=True)
    worksheet = workbook[workbook.sheetnames[0]]
    column_B = worksheet["B"]
    column_C = worksheet["C"]

    services_classifications = dict()
    try:
        for cell_B, cell_C in zip(column_B, column_C):
            b_value = cell_B.value
            c_value = cell_C.value
            if b_value and b_value.startswith("Serviço:"):
                service = b_value.split(": ", 1)[-1].upper()
                services_classifications.setdefault(service, [])
            elif not b_value and c_value:
                services_classifications[service].append(c_value.upper())
        workbook.close()
    except:
        raise DefaultException("non_standard_file", code=400)

    return services_classifications


def upload_services_and_classifications(file: FileStorage):

    file_extension = file.filename.split(".")[-1]

    if file_extension == "xlsx":
        services_classifications = read_xlsx(file=file)

    elif file_extension == "ods":
        services_classifications = read_ods(file=file)

    else:
        raise ValidationException(
            errors={"file": "The file type must be .xlsx or .ods"},
            message="invalid_type_file",
        )

    if not services_classifications:
        raise DefaultException("non_standard_file", code=400)

    try:

        for key, values in services_classifications.items():

            service_code, service_name = key.split(" - ", 1)

            if not check_numbers(service_code) and not service_name:
                raise DefaultException("non_standard_file", code=400)

            if service_exists(service_code=service_code, service_name=service_name):
                raise ValidationException(
                    errors={
                        f"key_in_use": f"The service with code='{service_code}' or name='{service_name}' already exists)"
                    },
                    message="unique_violation",
                )

            new_service = Service(code=service_code, name=service_name)

            db.session.add(new_service)
            db.session.flush()

            for value in values:

                classification_code, classification_name = value.split(" - ", 1)

                if not check_numbers(classification_code) and not classification_name:
                    raise DefaultException("non_standard_file", code=400)

                if _classification_exists(
                    service_id=new_service.id,
                    classification_code=classification_code,
                    classification_name=classification_name,
                ):
                    raise ValidationException(
                        errors={
                            f"key_in_use": f"The classification with pair (service_id={new_service.id} and code='{classification_code}') or (service_id={new_service.id} and name='{classification_name}') already exists)"
                        },
                        message="unique_violation",
                    )

                new_classification = Classification(
                    service_id=new_service.id,
                    code=classification_code,
                    name=classification_name,
                )

                db.session.add(new_classification)

            db.session.commit()

    except ValueError or IndexError:
        raise DefaultException("non_standard_file", code=400)


from .service_service import get_service, service_exists
